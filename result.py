# ライブラリのインポート
import cv2
import pyocr
from PIL import Image
import os
import requests
import sys
from datetime import datetime,date
import subprocess
import openpyxl as px
import time
from my_logger import set_logger, getLogger
from dotenv import load_dotenv
import csv

#start = time.time()  # 現在時刻（処理開始前）を取得

load_dotenv()
set_logger()
logger = getLogger(__name__)

# 環境設定の読み込み
SUMMARY_PY = os.getenv('SUMMARY_PY')
OUTPUT_DIR = os.getenv('OUTPUT_DIR')
TEMPLATE_DIR = os.getenv('TEMPLATE_DIR')
NAME_DICT_FLAG = os.getenv('NAME_DICT_FLAG')
NAME_DICT_CSV = os.getenv('NAME_DICT_CSV')

# Windows環境ならPATHを設定
if os.environ['PATH'][0] == "C":
    TESSERACT_PATH = os.getenv('TESSERACT_PATH')
    TESSDATA_PATH = os.getenv('TESSDATA_PATH')
    os.environ["PATH"] += os.pathsep + TESSERACT_PATH
    os.environ["TESSDATA_PREFIX"] = TESSDATA_PATH

# OCRエンジン取得
tools = pyocr.get_available_tools()
tool = tools[0]

# OCRの設定 ※tesseract_layout=6が精度には重要。デフォルトは3
builder = pyocr.builders.TextBuilder(tesseract_layout=6)

# データ補正（名前）
if int(NAME_DICT_FLAG):
    with open(NAME_DICT_CSV, encoding='utf8') as f:
        csvreader  = csv.DictReader(f,delimiter=",", doublequote=True, lineterminator="\r\n", quotechar='"', skipinitialspace=True)
        NAME_CORRECTIONS = [row for row in csvreader]

# 指定した画像をOCRしてテキストを抽出
def img2text(image_file, lang, place):

    # グレイスケール＆二値化
    img = cv2.imread(OUTPUT_DIR + image_file, 0); #グレイスケールで読み込む
    ret, img = cv2.threshold(img, 150, 255, cv2.THRESH_BINARY)
    cv2.imwrite(OUTPUT_DIR + str(place) + image_file, img)

    # OCR処理
    img = Image.open(OUTPUT_DIR + str(place) + image_file)
    txt = tool.image_to_string(img, lang=lang, builder=builder)
    txt = ''.join(txt.split())

    return(txt)

# 順位に応じたプレイヤーの名前、素点、順位点を出力
def player_info_print(template_file, place, flag, img):

    # テンプレート画像の読み込み
    template = cv2.imread(template_file)
    h, w, _ = template.shape # 幅と高さを取得

    # 画像の検索（Template Matching）
    result = cv2.matchTemplate(img, template, cv2.TM_CCORR_NORMED)

    # 検索結果の信頼度と位置座標の取得
    min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)

    # 信頼度が低い場合は終了
    CONFIDENCE_THRESHOLD = os.getenv('CONFIDENCE_THRESHOLD')
    if max_val < float(CONFIDENCE_THRESHOLD):
        logger.error(f"画像検索結果の信頼度({max_val})が閾値({CONFIDENCE_THRESHOLD})よりも低いため処理を終了します。")
        exit()

    # 検索結果の座標を取得
    top_left = max_loc # 左上
    bottom_right = (top_left[0] + w, top_left[1] + h) #右下

    # 検索対象画像内に、検索結果を水色の長方形で描画
    cv2.rectangle(img, top_left, bottom_right, (255, 255, 0), 2)

    # 名前と素点の座標を指定
    if flag == 'SP': #アプリ版
        if place == 1:
            player_name_tl  = (top_left[0] + 230, top_left[1] - 20)
            player_name_br  = (top_left[0] + 530, top_left[1] + 40)
            player_point_tl = (top_left[0] + 230, top_left[1] + 45)
            player_point_br = (top_left[0] + 530, top_left[1] + 135)
        else:
            player_name_tl  = (top_left[0] + 195, top_left[1] + 0)
            player_name_br  = (top_left[0] + 420, top_left[1] + 40)
            player_point_tl = (top_left[0] + 195, top_left[1] + 45)
            player_point_br = (top_left[0] + 420, top_left[1] + 115)
    else: #ブラウザ版
        if place == 1:
            player_name_tl  = (top_left[0] + 190, top_left[1] + 7)
            player_name_br  = (top_left[0] + 450, top_left[1] + 60)
            player_point_tl = (top_left[0] + 220, top_left[1] + 65)
            player_point_br = (top_left[0] + 450, top_left[1] + 120)
        else:
            player_name_tl  = (top_left[0] + 160, top_left[1] + 5)
            player_name_br  = (top_left[0] + 360, top_left[1] + 40)
            player_point_tl = (top_left[0] + 160, top_left[1] + 45)
            player_point_br = (top_left[0] + 360, top_left[1] + 95)

    # 指定した範囲を切り取り、画像ファイルとして保存
    cv2.imwrite(OUTPUT_DIR + 'player_name.png',  img[player_name_tl[1]:player_name_br[1],player_name_tl[0]:player_name_br[0]])
    cv2.imwrite(OUTPUT_DIR + 'player_point.png', img[player_point_tl[1]:player_point_br[1],player_point_tl[0]:player_point_br[0]])

    # 切り取った画像ファイルから文字認識
    name  = str(img2text('player_name.png', 'jpn', place))
    point = int(img2text('player_point.png','eng', place))

    # 順位点の計算
    uma = {1:15, 2:5, 3:-5, 4:-15} #順位ウマ
    rank = round(( point - 25000 ) / 1000 + uma[place], 2)

    # データ補正（名前）
    if int(NAME_DICT_FLAG):
        for wrong, correct in NAME_CORRECTIONS[0].items():
            name = name.replace(str(wrong), str(correct))

    # 1位の場合のみヘッダを標準出力
    if place == 1:
        print("==========対局結果==========")

    # 名前、素点、順位点を整形して標準出力（名前が短い人はタブを1つ追加）
    name_padding = '\t' if len(name.encode('utf-8')) < 17 else ''
    print(f"{place: >2}  {name}{name_padding}\t{point: >6}\t{round(float(rank),2): >5}")

    # 集計用Excelファイルに追記
    BIKOU = os.getenv('BIKOU')
    updatelist(today, now, place, name, point, rank, str(BIKOU))

    # 検索対象画像内に、検索結果を黄色い長方形で描画
    cv2.rectangle(img, player_name_tl, player_name_br, (0, 255, 255), 2)
    cv2.rectangle(img, player_point_tl, player_point_br, (0, 255, 255), 2)

# 集計用Excelに対局結果を追記
def updatelist(today, now, place, name, point, rank, comment):
    
    # 集計用Excelファイル
    RESULT_EXCEL = os.getenv('RESULT_EXCEL')
    sheet_name = '対局結果'

    # Excelファイルを開く
    wb = px.load_workbook(RESULT_EXCEL)
    ws = wb[sheet_name]

    # 最大行番号を取得
    max = ws.max_row

    # 最終行に結果を追記
    ws.cell(row=max+1, column=1).value = today
    ws.cell(row=max+1, column=2).value = now
    ws.cell(row=max+1, column=3).value = place
    ws.cell(row=max+1, column=4).value = name
    ws.cell(row=max+1, column=5).value = point
    ws.cell(row=max+1, column=6).value = rank
    ws.cell(row=max+1, column=7).value = comment

    # Excelファイルを上書き保存
    wb.save(RESULT_EXCEL)

##############
#    MAIN    #
##############

# 検索対象画像をダウンロードして保存
#url = input("画像のURLを入力してください: ")
INPUT_IMAGE = os.getenv('INPUT_IMAGE')

# 引数(URL)があればダウンロードする
if len(sys.argv) > 1:
    urlData = requests.get(sys.argv[1]).content
    with open(INPUT_IMAGE, "wb") as aaa:
        aaa.write(urlData)

# 画像をOpenCVで読み込む
try:
    img = cv2.imread(INPUT_IMAGE)
    if img is None:
        raise ValueError(f"画像 '{INPUT_IMAGE}' を読み込めませんでした。")
except Exception as ex:
    print(f"エラー発生: {ex}")
    logger.error(f"エラー発生: {ex}")
    sys.exit()

# 画像のサイズからPCかSPかを決める
INPUT_IMAGE_WIDTH = os.getenv('INPUT_IMAGE_WIDTH')
if img.shape[1] < int(INPUT_IMAGE_WIDTH):
    flag = 'PC' #ブラウザ版
else:
    flag = 'SP' #アプリ版

# 現在日時を取得
today = date.today()
now = datetime.now().time()

# 順位ごとに処理
player_info_print(TEMPLATE_DIR + 'no1.png', 1, flag, img)
player_info_print(TEMPLATE_DIR + 'no2.png', 2, flag, img)
player_info_print(TEMPLATE_DIR + 'no3.png', 3, flag, img)
player_info_print(TEMPLATE_DIR + 'no4.png', 4, flag, img)

# 検索結果を描画した画像を出力
WRITE_IMAGE = os.getenv('WRITE_IMAGE')
cv2.imwrite(WRITE_IMAGE, img)

# 集計用コードを実行
command = ['python', SUMMARY_PY] 
proc = subprocess.run(command, capture_output=True, text=True).stdout
print(proc)

# 集計用Excelのリンクを出力
print("集計用Excelファイルはこちら")
print(os.getenv('SUMMARY_EXCEL'))

#end = time.time()  # 現在時刻（処理完了後）を取得
#print(f"処理時間: {round(end - start, 2)}秒")  # 処理にかかった時間データを使用
